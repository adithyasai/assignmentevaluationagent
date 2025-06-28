"""
Excel file processing utilities for the Assignment Agent.
"""
import pandas as pd
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from loguru import logger
import config


class ExcelHandler:
    """Handles Excel file operations for student data and grading results."""
    
    def __init__(self):
        """Initialize the Excel handler."""
        self.students_data: Optional[pd.DataFrame] = None
        self.original_file_path: Optional[str] = None
    
    def load_students_file(self, file_path: str) -> Tuple[bool, str]:
        """
        Load student data from Excel file.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Tuple of (success, message)
        """
        try:
            self.students_data = pd.read_excel(file_path)
            self.original_file_path = file_path
            
            # Add output columns if they don't exist
            output_cols = config.EXCEL_COLUMNS['OUTPUT']
            for col_name in output_cols.values():
                if col_name not in self.students_data.columns:
                    if col_name == output_cols['BUILD_STATUS']:
                        self.students_data[col_name] = config.BUILD_STATUS['PENDING']
                    elif col_name == output_cols['GRADE']:
                        self.students_data[col_name] = 0
                    elif col_name == output_cols['PROCESSED_AT']:
                        self.students_data[col_name] = ''
                    else:
                        self.students_data[col_name] = ''
            
            logger.info(f"Loaded {len(self.students_data)} student records from {file_path}")
            return True, f"Successfully loaded {len(self.students_data)} student records"
            
        except Exception as e:
            logger.error(f"Failed to load Excel file: {str(e)}")
            return False, f"Failed to load Excel file: {str(e)}"
    
    def get_students_data(self) -> Optional[pd.DataFrame]:
        """
        Get the loaded students data.
        
        Returns:
            DataFrame with student data or None if not loaded
        """
        return self.students_data.copy() if self.students_data is not None else None
    
    def get_student_info(self, index: int) -> Optional[Dict[str, str]]:
        """
        Get information for a specific student.
        
        Args:
            index: Row index of the student
            
        Returns:
            Dictionary with student information or None if not found
        """
        if self.students_data is None or index >= len(self.students_data):
            return None
        
        try:
            row = self.students_data.iloc[index]
            required_cols = config.EXCEL_COLUMNS['REQUIRED']
            optional_cols = config.EXCEL_COLUMNS['OPTIONAL']
            
            student_info = {
                'name': str(row[required_cols['NAME']]),
                'github_url': str(row[required_cols['GITHUB_URL']]),
                'index': index
            }
            
            # Add optional fields if they exist and are not empty
            for key, col_name in optional_cols.items():
                if col_name in row and pd.notna(row[col_name]):
                    student_info[key.lower()] = str(row[col_name])
            
            return student_info
            
        except Exception as e:
            logger.error(f"Failed to get student info for index {index}: {str(e)}")
            return None
    
    def update_student_result(self, index: int, build_status: str, grade: int, 
                            feedback: str = "", build_errors: str = "") -> bool:
        """
        Update grading results for a specific student.
        
        Args:
            index: Row index of the student
            build_status: Build status (Success/Failed/Error)
            grade: Numerical grade (0-100)
            feedback: Detailed feedback text
            build_errors: Technical error messages
            
        Returns:
            True if update successful, False otherwise
        """
        if self.students_data is None or index >= len(self.students_data):
            return False
        
        try:
            output_cols = config.EXCEL_COLUMNS['OUTPUT']
            
            self.students_data.loc[index, output_cols['BUILD_STATUS']] = build_status
            self.students_data.loc[index, output_cols['GRADE']] = grade
            self.students_data.loc[index, output_cols['FEEDBACK']] = feedback
            self.students_data.loc[index, output_cols['BUILD_ERRORS']] = build_errors
            self.students_data.loc[index, output_cols['PROCESSED_AT']] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            logger.debug(f"Updated results for student at index {index}: {build_status}, Grade: {grade}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to update student result for index {index}: {str(e)}")
            return False
    
    def save_results(self, output_path: Optional[str] = None) -> Tuple[bool, str]:
        """
        Save the updated results to an Excel file.
        
        Args:
            output_path: Path for output file. If None, generates timestamped filename
            
        Returns:
            Tuple of (success, file_path or error_message)
        """
        if self.students_data is None:
            return False, "No data to save"
        
        try:
            if output_path is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = config.OUTPUT_FILE_PATTERNS['GRADED_EXCEL'].format(timestamp=timestamp)
                output_path = str(config.DATA_DIR / filename)
            
            # Create output directory if it doesn't exist
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            
            # Save to Excel with formatting
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                self.students_data.to_excel(writer, sheet_name='Grading Results', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Grading Results']
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add color coding for build status
                from openpyxl.styles import PatternFill
                
                build_status_col = None
                for idx, col in enumerate(self.students_data.columns):
                    if col == config.EXCEL_COLUMNS['OUTPUT']['BUILD_STATUS']:
                        build_status_col = idx + 1
                        break
                
                if build_status_col:
                    success_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                    failed_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                    warning_fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
                    
                    for row in range(2, len(self.students_data) + 2):  # Skip header row
                        cell = worksheet.cell(row=row, column=build_status_col)
                        status = str(cell.value).lower()
                        
                        if 'success' in status:
                            cell.fill = success_fill
                        elif 'failed' in status or 'error' in status:
                            cell.fill = failed_fill
                        elif 'warning' in status:
                            cell.fill = warning_fill
            
            logger.info(f"Results saved to {output_path}")
            return True, output_path
            
        except Exception as e:
            logger.error(f"Failed to save results: {str(e)}")
            return False, f"Failed to save results: {str(e)}"
    
    def get_summary_stats(self) -> Dict[str, int]:
        """
        Get summary statistics of grading results.
        
        Returns:
            Dictionary with summary statistics
        """
        if self.students_data is None:
            return {}
        
        try:
            output_cols = config.EXCEL_COLUMNS['OUTPUT']
            build_status_col = output_cols['BUILD_STATUS']
            grade_col = output_cols['GRADE']
            
            stats = {
                'total_students': len(self.students_data),
                'processed': len(self.students_data[self.students_data[build_status_col] != config.BUILD_STATUS['PENDING']]),
                'success': len(self.students_data[self.students_data[build_status_col] == config.BUILD_STATUS['SUCCESS']]),
                'failed': len(self.students_data[self.students_data[build_status_col] == config.BUILD_STATUS['FAILED']]),
                'errors': len(self.students_data[self.students_data[build_status_col] == config.BUILD_STATUS['ERROR']]),
                'average_grade': 0,
                'max_grade': 0,
                'min_grade': 0
            }
            
            # Calculate grade statistics for processed students
            processed_grades = self.students_data[
                self.students_data[build_status_col] != config.BUILD_STATUS['PENDING']
            ][grade_col]
            
            if len(processed_grades) > 0:
                stats['average_grade'] = round(processed_grades.mean(), 2)
                stats['max_grade'] = int(processed_grades.max())
                stats['min_grade'] = int(processed_grades.min())
            
            return stats
            
        except Exception as e:
            logger.error(f"Failed to calculate summary stats: {str(e)}")
            return {}
    
    def export_error_log(self, output_path: Optional[str] = None) -> Tuple[bool, str]:
        """
        Export detailed error log for failed builds.
        
        Args:
            output_path: Path for output file. If None, generates timestamped filename
            
        Returns:
            Tuple of (success, file_path or error_message)
        """
        if self.students_data is None:
            return False, "No data available"
        
        try:
            if output_path is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = config.OUTPUT_FILE_PATTERNS['ERROR_LOG'].format(timestamp=timestamp)
                output_path = str(config.LOGS_DIR / filename)
            
            # Create output directory if it doesn't exist
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            
            output_cols = config.EXCEL_COLUMNS['OUTPUT']
            required_cols = config.EXCEL_COLUMNS['REQUIRED']
            
            # Filter for failed/error records with build errors
            failed_records = self.students_data[
                (self.students_data[output_cols['BUILD_STATUS']].isin([
                    config.BUILD_STATUS['FAILED'], 
                    config.BUILD_STATUS['ERROR']
                ])) &
                (self.students_data[output_cols['BUILD_ERRORS']].notna()) &
                (self.students_data[output_cols['BUILD_ERRORS']] != '')
            ]
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(f"Assignment Grading Error Log\n")
                f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Total Failed/Error Records: {len(failed_records)}\n")
                f.write("=" * 80 + "\n\n")
                
                for idx, row in failed_records.iterrows():
                    f.write(f"Student: {row[required_cols['NAME']]}\n")
                    f.write(f"GitHub URL: {row[required_cols['GITHUB_URL']]}\n")
                    f.write(f"Status: {row[output_cols['BUILD_STATUS']]}\n")
                    f.write(f"Grade: {row[output_cols['GRADE']]}\n")
                    f.write(f"Processed At: {row[output_cols['PROCESSED_AT']]}\n")
                    f.write(f"Feedback: {row[output_cols['FEEDBACK']]}\n")
                    f.write(f"Build Errors:\n{row[output_cols['BUILD_ERRORS']]}\n")
                    f.write("-" * 80 + "\n\n")
            
            logger.info(f"Error log exported to {output_path}")
            return True, output_path
            
        except Exception as e:
            logger.error(f"Failed to export error log: {str(e)}")
            return False, f"Failed to export error log: {str(e)}"
